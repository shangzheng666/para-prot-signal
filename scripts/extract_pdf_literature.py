"""Extract structured metadata from PDF literature via DeepSeek API.

Walks a directory of PDFs, extracts text, asks DeepSeek to populate a fixed
schema for each paper, and writes the result to an Excel sheet plus a
processing log.

Usage:
  python scripts/extract_pdf_literature.py \
      --pdf-dir pdfs \
      --out extraction_summary.xlsx \
      --log extraction_log.txt

Requires:
  pip install pdfplumber openpyxl
  DEEPSEEK_API_KEY env var (or --api-key)
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path


DEEPSEEK_URL = "https://api.deepseek.com/v1/chat/completions"
DEEPSEEK_MODEL = "deepseek-chat"

# Fields produced by the LLM. The Excel "filename" column comes from the OS.
LLM_FIELDS = [
  "first_author",
  "corresponding_author",
  "year",
  "journal",
  "GBP1_species",
  "has_14_3_3",
  "phospho_sites",
  "interaction_method",
  "key_conclusion",
  "toxoplasma_neospora_model",
]

EXCEL_HEADERS = ["filename", *LLM_FIELDS]

SYSTEM_PROMPT = (
  "You extract structured metadata from scientific papers about GBP1, "
  "14-3-3 proteins, and Toxoplasma/Neospora research. "
  "You return ONLY a JSON object. You never invent facts."
)

USER_PROMPT_TEMPLATE = """Extract metadata from the paper text below.

STRICT RULES:
1. Use ONLY information explicitly stated in the text. If something is not
   stated, set the field to "N/A". Do not infer, summarize, or fabricate.
2. phospho_sites: quote the EXACT residue notation as it appears in the text
   (e.g. "Ser156", "Thr99", "S156/T172"). If multiple sites are mentioned,
   join them with "; ". If none are mentioned, "N/A".
3. key_conclusion: copy ONE sentence VERBATIM from the paper (English) that
   best states the main finding. Do not paraphrase. Surround with no quotes.
4. has_14_3_3: "Yes" only if 14-3-3 is mentioned. If an isoform is named,
   write "Yes (14-3-3X)" where X is the isoform letter/number. Else "No".
5. GBP1_species: e.g. "human GBP1", "mouse Gbp2 paralog". Use the exact
   organism wording from the paper.
6. interaction_method: list the experimental methods used to validate
   interactions, separated by "; " (e.g. "co-IP; pull-down; ITC").
7. toxoplasma_neospora_model: "Yes" with a short note if T. gondii or
   N. caninum infection is used as a model. Else "No".
8. year: 4-digit publication year only.

Return ONLY this JSON object (no markdown, no prose), with all keys present:
{{
  "first_author": "",
  "corresponding_author": "",
  "year": "",
  "journal": "",
  "GBP1_species": "",
  "has_14_3_3": "",
  "phospho_sites": "",
  "interaction_method": "",
  "key_conclusion": "",
  "toxoplasma_neospora_model": ""
}}

Paper text:
<<<
{text}
>>>
"""


def extract_pdf_text(pdf_path: Path) -> tuple[str, str]:
  """Return (text, note). note is empty on clean success, else a warning."""
  try:
    import pdfplumber
  except ImportError:
    raise SystemExit(
      "pdfplumber is not installed. Run: pip install pdfplumber openpyxl"
    )

  pages: list[str] = []
  try:
    with pdfplumber.open(pdf_path) as pdf:
      for page in pdf.pages:
        text = page.extract_text() or ""
        if text:
          pages.append(text)
  except Exception as exc:  # noqa: BLE001 — pdfplumber raises many types
    return "", f"pdf_open_failed: {type(exc).__name__}: {exc}"

  joined = "\n".join(pages)
  joined = re.sub(r"[ \t]+", " ", joined)
  joined = re.sub(r"\n{3,}", "\n\n", joined).strip()

  if len(joined) < 500:
    return joined, "low_text_yield_likely_scanned_pdf_needs_ocr"
  return joined, ""


def truncate_text(text: str, max_chars: int) -> str:
  """Keep head + tail when text exceeds max_chars (results often sit at end)."""
  if len(text) <= max_chars:
    return text
  head = int(max_chars * 0.75)
  tail = max_chars - head
  return text[:head] + "\n\n[... truncated ...]\n\n" + text[-tail:]


def call_deepseek(api_key: str, paper_text: str, *, timeout: int = 120) -> str:
  """POST to DeepSeek with retry on 5xx/network errors. Returns raw JSON string."""
  payload = {
    "model": DEEPSEEK_MODEL,
    "messages": [
      {"role": "system", "content": SYSTEM_PROMPT},
      {"role": "user", "content": USER_PROMPT_TEMPLATE.format(text=paper_text)},
    ],
    "response_format": {"type": "json_object"},
    "temperature": 0.0,
    "max_tokens": 1024,
  }
  data = json.dumps(payload).encode("utf-8")
  headers = {
    "Authorization": f"Bearer {api_key}",
    "Content-Type": "application/json",
  }

  max_attempts = 3
  backoff = 2
  for attempt in range(1, max_attempts + 1):
    req = urllib.request.Request(DEEPSEEK_URL, data=data, headers=headers, method="POST")
    try:
      with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = json.loads(resp.read().decode("utf-8"))
      return body["choices"][0]["message"]["content"]
    except urllib.error.HTTPError as exc:
      if exc.code == 429 or exc.code >= 500:
        if attempt == max_attempts:
          raise
        print(
          f"[retry] DeepSeek HTTP {exc.code} on attempt {attempt}/{max_attempts}, "
          f"retrying in {backoff}s ...",
          file=sys.stderr,
        )
      else:
        raise  # 4xx other than 429 — surface immediately
    except (urllib.error.URLError, TimeoutError, ConnectionError) as exc:
      if attempt == max_attempts:
        raise
      print(
        f"[retry] {type(exc).__name__} on attempt {attempt}/{max_attempts}, "
        f"retrying in {backoff}s ...",
        file=sys.stderr,
      )
    time.sleep(backoff)
    backoff *= 2

  raise RuntimeError("call_deepseek: exhausted retries")


def parse_llm_json(raw: str) -> dict[str, str]:
  """Parse the LLM JSON and ensure all fields are present as strings."""
  obj = json.loads(raw)
  out: dict[str, str] = {}
  for key in LLM_FIELDS:
    value = obj.get(key, "")
    if value is None:
      value = ""
    if not isinstance(value, str):
      value = str(value)
    out[key] = value.strip() or "N/A"
  return out


def empty_row() -> dict[str, str]:
  return {key: "N/A" for key in LLM_FIELDS}


def count_missing(row: dict[str, str]) -> int:
  return sum(1 for k in LLM_FIELDS if row.get(k, "N/A") == "N/A")


def write_excel(rows: list[dict[str, str]], out_path: Path) -> None:
  try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
  except ImportError:
    raise SystemExit(
      "openpyxl is not installed. Run: pip install pdfplumber openpyxl"
    )

  wb = Workbook()
  ws = wb.active
  ws.title = "extraction"
  ws.append(EXCEL_HEADERS)
  for cell in ws[1]:
    cell.font = Font(bold=True)

  for row in rows:
    ws.append([row.get(h, "") for h in EXCEL_HEADERS])

  for idx, header in enumerate(EXCEL_HEADERS, start=1):
    width = 18
    if header in {"key_conclusion", "phospho_sites", "interaction_method"}:
      width = 60
    elif header == "filename":
      width = 40
    ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

  wb.save(out_path)


def process_pdf(
  pdf_path: Path,
  api_key: str,
  max_chars: int,
  dry_run: bool,
) -> tuple[dict[str, str], str]:
  """Return (row, status_note). Row always includes filename."""
  text, warn = extract_pdf_text(pdf_path)
  row: dict[str, str] = {"filename": pdf_path.name, **empty_row()}

  if not text:
    return row, f"failed: {warn or 'no_text_extracted'}"

  notes: list[str] = []
  if warn:
    notes.append(warn)

  paper_text = truncate_text(text, max_chars)

  if dry_run:
    notes.append("dry_run_no_api_call")
    return row, "; ".join(notes) if notes else "dry_run"

  try:
    raw = call_deepseek(api_key, paper_text)
  except Exception as exc:  # noqa: BLE001
    notes.append(f"api_error: {type(exc).__name__}: {exc}")
    return row, "failed: " + "; ".join(notes)

  try:
    parsed = parse_llm_json(raw)
  except json.JSONDecodeError as exc:
    notes.append(f"json_parse_error: {exc}")
    notes.append(f"raw_head: {raw[:200]!r}")
    return row, "failed: " + "; ".join(notes)

  row.update(parsed)
  missing = count_missing(row)
  if missing >= len(LLM_FIELDS) // 2:
    notes.append(f"partial_missing_{missing}_of_{len(LLM_FIELDS)}")
    status = "partial: " + "; ".join(notes) if notes else "partial"
  else:
    status = "success" if not notes else "success_with_warnings: " + "; ".join(notes)
  return row, status


def main() -> int:
  parser = argparse.ArgumentParser(
    description="Extract structured metadata from PDF literature via DeepSeek."
  )
  parser.add_argument("--pdf-dir", default="pdfs", help="Directory containing PDFs.")
  parser.add_argument(
    "--out", default="extraction_summary.xlsx", help="Output Excel path."
  )
  parser.add_argument(
    "--log", default="extraction_log.txt", help="Output log path."
  )
  parser.add_argument(
    "--api-key", default=None, help="DeepSeek API key (or set DEEPSEEK_API_KEY)."
  )
  parser.add_argument(
    "--max-chars",
    type=int,
    default=30000,
    help="Max characters of paper text sent to the API (default 30000).",
  )
  parser.add_argument(
    "--limit",
    type=int,
    default=0,
    help="Process at most N PDFs (0 = no limit). Useful for a first dry run.",
  )
  parser.add_argument(
    "--dry-run",
    action="store_true",
    help="Skip API calls; only verify PDF text extraction.",
  )
  args = parser.parse_args()

  pdf_dir = Path(args.pdf_dir)
  if not pdf_dir.is_dir():
    print(f"error: --pdf-dir {pdf_dir} is not a directory", file=sys.stderr)
    return 2

  pdfs = sorted(p for p in pdf_dir.iterdir() if p.suffix.lower() == ".pdf")
  if args.limit > 0:
    pdfs = pdfs[: args.limit]
  if not pdfs:
    print(f"error: no PDFs found in {pdf_dir}", file=sys.stderr)
    return 2

  api_key = args.api_key or os.environ.get("DEEPSEEK_API_KEY", "")
  if not api_key and not args.dry_run:
    print(
      "error: DEEPSEEK_API_KEY not set. Pass --api-key or set the env var, "
      "or use --dry-run to test PDF extraction only.",
      file=sys.stderr,
    )
    return 2

  print(f"Found {len(pdfs)} PDF(s) in {pdf_dir}", file=sys.stderr)
  if args.dry_run:
    print("Dry run: skipping DeepSeek API calls.", file=sys.stderr)

  rows: list[dict[str, str]] = []
  log_lines: list[str] = []
  field_miss_counter: dict[str, int] = {f: 0 for f in LLM_FIELDS}
  success = partial = failed = 0

  for i, pdf in enumerate(pdfs, 1):
    print(f"[{i}/{len(pdfs)}] {pdf.name}", file=sys.stderr)
    started = time.time()
    row, status = process_pdf(pdf, api_key, args.max_chars, args.dry_run)
    elapsed = time.time() - started

    for f in LLM_FIELDS:
      if row.get(f, "N/A") == "N/A":
        field_miss_counter[f] += 1

    if status.startswith("success"):
      success += 1
    elif status.startswith("partial"):
      partial += 1
    else:
      failed += 1

    rows.append(row)
    ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_lines.append(f"[{ts}] ({elapsed:5.1f}s) {pdf.name} | {status}")
    print(
      f"      -> {status[:120]}{'...' if len(status) > 120 else ''}",
      file=sys.stderr,
    )

  out_xlsx = Path(args.out)
  out_log = Path(args.log)
  write_excel(rows, out_xlsx)
  out_log.write_text("\n".join(log_lines) + "\n", encoding="utf-8")

  print("", file=sys.stderr)
  print(f"Wrote {len(rows)} rows to {out_xlsx}", file=sys.stderr)
  print(f"Wrote log to {out_log}", file=sys.stderr)
  print(
    f"Summary: success={success}  partial={partial}  failed={failed}",
    file=sys.stderr,
  )

  top_missing = sorted(field_miss_counter.items(), key=lambda kv: -kv[1])[:3]
  if any(c > 0 for _, c in top_missing):
    print("Most-missing fields (review these manually):", file=sys.stderr)
    for f, c in top_missing:
      if c > 0:
        print(f"  - {f}: {c}/{len(rows)} missing", file=sys.stderr)

  return 0


if __name__ == "__main__":
  raise SystemExit(main())
